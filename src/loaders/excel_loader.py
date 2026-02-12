from pathlib import Path
from typing import Optional, Literal
from openpyxl import load_workbook

from src.loaders.base_loader import BaseLoader
from src.utils.document import Document

"""
ExcelLoader - Load Microsoft Excel spreadsheets
Extracts data from .xlsx files, converting sheets to documents
"""
"""
Load a Microsoft Excel (.xlsx) file and extract content
Can extract data as:
    - One document per sheet
    - One document per row
    - Single document for entire workbook
Attributes:
    file_path: Path to the Excel file
    mode: Extraction mode ('sheet', 'row', or 'workbook')
Example:
    >>> loader = ExcelLoader("data.xlsx", mode="sheet")
    >>> docs = loader.load()
    >>> print(f"Loaded {len(docs)} sheets")
"""
class ExcelLoader(BaseLoader):
    

    """
    Initialize the ExcelLoader
    Args:
        file_path: Path to the Excel file
        mode: How to split content:
                'sheet' - one document per sheet
                'row' - one document per row
                'workbook' - single document for all
        sheet_names: Specific sheets to load (None = all)
        metadata: Optional additional metadata
    """
    def __init__(
        self,
        file_path: str,
        mode: Literal["sheet", "row", "workbook"] = "sheet",
        sheet_names: Optional[list[str]] = None,
        metadata: Optional[dict] = None
    ):
        self.file_path = Path(file_path)
        self.mode = mode
        self.sheet_names = sheet_names
        self.extra_metadata = metadata or {}
    
    """
    Load the Excel file and extract content
    Returns:
        List of Documents based on the extraction mode
    Raises:
        FileNotFoundError: If the file doesn't exist
    """
    def load(self) -> list[Document]:
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")
        
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        
        """
        Determine which sheets to process
        """
        sheets_to_process = self.sheet_names or wb.sheetnames
        
        documents = []
        
        """
        Load based on mode
        """
        if self.mode == "workbook":
            documents = self._load_as_workbook(wb, sheets_to_process)
        elif self.mode == "sheet":
            documents = self._load_as_sheets(wb, sheets_to_process)
        elif self.mode == "row":
            documents = self._load_as_rows(wb, sheets_to_process)
        
        wb.close()
        return documents
    
    """
    Load entire workbook as a single document
    """
    def _load_as_workbook(self, wb, sheets: list[str]) -> list[Document]:
        all_content = []
        
        for sheet_name in sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_content = self._sheet_to_text(ws, sheet_name)
                all_content.append(sheet_content)
        
        content = "\n\n---\n\n".join(all_content)
        
        metadata = {
            "source": str(self.file_path),
            "file_name": self.file_path.name,
            "file_type": "xlsx",
            "sheet_count": len(sheets),
            **self.extra_metadata
        }
        
        return [Document(page_content=content, metadata=metadata)]
    
    """
    Load each sheet as a separate document
    """
    def _load_as_sheets(self, wb, sheets: list[str]) -> list[Document]:
        documents = []
        
        for sheet_name in sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content = self._sheet_to_text(ws, sheet_name)
                
                metadata = {
                    "source": str(self.file_path),
                    "file_name": self.file_path.name,
                    "file_type": "xlsx",
                    "sheet_name": sheet_name,
                    **self.extra_metadata
                }
                
                if content.strip():
                    documents.append(Document(page_content=content, metadata=metadata))
        
        return documents
    
    """
    Load each row as a separate document
    """
    def _load_as_rows(self, wb, sheets: list[str]) -> list[Document]:
        documents = []
        
        for sheet_name in sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                if not rows:
                    continue
                
                # Use first row as headers
                headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
                
                for row_idx, row in enumerate(rows[1:], start=2):
                    row_data = []
                    for header, value in zip(headers, row):
                        if value is not None:
                            row_data.append(f"{header}: {value}")
                    
                    if row_data:
                        content = "\n".join(row_data)
                        
                        metadata = {
                            "source": str(self.file_path),
                            "file_name": self.file_path.name,
                            "file_type": "xlsx",
                            "sheet_name": sheet_name,
                            "row": row_idx,
                            **self.extra_metadata
                        }
                        
                        documents.append(Document(page_content=content, metadata=metadata))
        
        return documents
    
    """
    Convert a worksheet to text format
    """
    def _sheet_to_text(self, ws, sheet_name: str) -> str:
        lines = [f"# {sheet_name}\n"]
        
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(row_values):  # Skip empty rows
                lines.append(" | ".join(row_values))
        
        return "\n".join(lines)
